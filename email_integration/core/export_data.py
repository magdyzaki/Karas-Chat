"""
نظام تصدير البيانات
Data Export System for EFM
"""
import csv
import os
from datetime import datetime
from typing import List, Dict, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not be available.")

from .db import (
    get_connection,
    get_all_clients,
    get_client_messages,
    get_clients_needing_followup
)


def export_clients_to_csv(file_path: str, clients: Optional[List] = None) -> bool:
    """
    تصدير قائمة العملاء إلى ملف CSV
    
    Args:
        file_path: مسار ملف CSV
        clients: قائمة العملاء (إذا كانت None، سيتم جلب جميع العملاء)
    
    Returns:
        True في حالة النجاح، False في حالة الفشل
    """
    try:
        if clients is None:
            clients = get_all_clients()
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # رأس الجدول
            writer.writerow([
                'ID', 'Company Name', 'Country', 'Contact Person',
                'Email', 'Phone', 'Website', 'Date Added',
                'Status', 'Score', 'Classification', 'Focus'
            ])
            
            # البيانات
            for client in clients:
                (
                    client_id, company, country, contact, email,
                    phone, website, date_added, status, score,
                    classification, is_focus
                ) = client
                
                writer.writerow([
                    client_id, company or '', country or '', contact or '',
                    email or '', phone or '', website or '', date_added or '',
                    status or '', score or 0, classification or '', 'Yes' if is_focus else 'No'
                ])
        
        return True
        
    except Exception as e:
        print(f"خطأ في تصدير CSV: {e}")
        return False


def export_clients_to_excel(file_path: str, clients: Optional[List] = None) -> bool:
    """
    تصدير قائمة العملاء إلى ملف Excel
    
    Args:
        file_path: مسار ملف Excel (.xlsx)
        clients: قائمة العملاء (إذا كانت None، سيتم جلب جميع العملاء)
    
    Returns:
        True في حالة النجاح، False في حالة الفشل
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install it using: pip install openpyxl")
    
    try:
        if clients is None:
            clients = get_all_clients()
        
        # إنشاء workbook جديد
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clients"
        
        # تنسيق الرأس
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # رأس الجدول
        headers = [
            'ID', 'Company Name', 'Country', 'Contact Person',
            'Email', 'Phone', 'Website', 'Date Added',
            'Status', 'Score', 'Classification', 'Focus'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # البيانات مع تنظيف وتحسين
        data_font = Font(size=10, name="Arial")
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        def clean_value(value):
            """تنظيف القيم من الأحرف الخاصة التي قد تسبب مشاكل في Excel"""
            if value is None:
                return ''
            value_str = str(value).strip()
            # إزالة الأحرف الخاصة التي قد تسبب مشاكل في Excel
            value_str = value_str.replace('\x00', '').replace('\r', ' ').replace('\n', ' ')
            # إزالة أحرف التحكم الأخرى
            value_str = ''.join(char for char in value_str if ord(char) >= 32 or char in '\t')
            return value_str
        
        for row_num, client in enumerate(clients, 2):
            (
                client_id, company, country, contact, email,
                phone, website, date_added, status, score,
                classification, is_focus
            ) = client
            
            ws.cell(row=row_num, column=1, value=clean_value(client_id))
            ws.cell(row=row_num, column=2, value=clean_value(company))
            ws.cell(row=row_num, column=3, value=clean_value(country))
            ws.cell(row=row_num, column=4, value=clean_value(contact))
            ws.cell(row=row_num, column=5, value=clean_value(email))
            ws.cell(row=row_num, column=6, value=clean_value(phone))
            ws.cell(row=row_num, column=7, value=clean_value(website))
            ws.cell(row=row_num, column=8, value=clean_value(date_added))
            ws.cell(row=row_num, column=9, value=clean_value(status))
            ws.cell(row=row_num, column=10, value=score if score else 0)
            ws.cell(row=row_num, column=11, value=clean_value(classification))
            ws.cell(row=row_num, column=12, value='Yes' if is_focus else 'No')
            
            # تطبيق التنسيق على الصف
            for col_num in range(1, 13):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
        
        # ضبط عرض الأعمدة (محسّن)
        column_widths = [8, 35, 18, 22, 30, 20, 35, 14, 12, 10, 18, 10]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # تجميد الصف الأول
        ws.freeze_panes = 'A2'
        
        # حفظ الملف
        wb.save(file_path)
        return True
        
    except Exception as e:
        print(f"خطأ في تصدير Excel: {e}")
        return False


def export_messages_to_csv(file_path: str, client_id: Optional[int] = None) -> bool:
    """
    تصدير الرسائل إلى ملف CSV
    
    Args:
        file_path: مسار ملف CSV
        client_id: معرف العميل (إذا كان None، سيتم تصدير جميع الرسائل)
    
    Returns:
        True في حالة النجاح، False في حالة الفشل
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        if client_id:
            query = """
                SELECT 
                    m.id, c.company_name, m.message_date, m.message_type,
                    m.channel, m.client_response, m.score_effect
                FROM messages m
                JOIN clients c ON m.client_id = c.id
                WHERE m.client_id = ?
                ORDER BY m.message_date DESC
            """
            cur.execute(query, (client_id,))
        else:
            query = """
                SELECT 
                    m.id, c.company_name, m.message_date, m.message_type,
                    m.channel, m.client_response, m.score_effect
                FROM messages m
                JOIN clients c ON m.client_id = c.id
                ORDER BY m.message_date DESC
            """
            cur.execute(query)
        
        messages = cur.fetchall()
        conn.close()
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # رأس الجدول
            writer.writerow([
                'ID', 'Client', 'Date', 'Type', 'Channel', 'Subject', 'Score Effect'
            ])
            
            # البيانات
            for msg in messages:
                msg_id, company, msg_date, msg_type, channel, subject, score_effect = msg
                
                writer.writerow([
                    msg_id, company or '', msg_date or '', msg_type or '',
                    channel or '', subject or '', score_effect or 0
                ])
        
        return True
        
    except Exception as e:
        print(f"خطأ في تصدير الرسائل CSV: {e}")
        return False


def export_messages_to_excel(file_path: str, client_id: Optional[int] = None) -> bool:
    """
    تصدير الرسائل إلى ملف Excel
    
    Args:
        file_path: مسار ملف Excel (.xlsx)
        client_id: معرف العميل (إذا كان None، سيتم تصدير جميع الرسائل)
    
    Returns:
        True في حالة النجاح، False في حالة الفشل
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install it using: pip install openpyxl")
    
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        if client_id:
            query = """
                SELECT 
                    m.id, c.company_name, m.message_date, m.message_type,
                    m.channel, m.client_response, m.score_effect
                FROM messages m
                JOIN clients c ON m.client_id = c.id
                WHERE m.client_id = ?
                ORDER BY m.message_date DESC
            """
            cur.execute(query, (client_id,))
        else:
            query = """
                SELECT 
                    m.id, c.company_name, m.message_date, m.message_type,
                    m.channel, m.client_response, m.score_effect
                FROM messages m
                JOIN clients c ON m.client_id = c.id
                ORDER BY m.message_date DESC
            """
            cur.execute(query)
        
        messages = cur.fetchall()
        conn.close()
        
        # إنشاء workbook جديد
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Messages"
        
        # تنسيق الرأس
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # رأس الجدول
        headers = ['ID', 'Client', 'Date', 'Type', 'Channel', 'Subject', 'Score Effect']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # البيانات
        for row_num, msg in enumerate(messages, 2):
            msg_id, company, msg_date, msg_type, channel, subject, score_effect = msg
            
            ws.cell(row=row_num, column=1, value=msg_id)
            ws.cell(row=row_num, column=2, value=company or '')
            ws.cell(row=row_num, column=3, value=msg_date or '')
            ws.cell(row=row_num, column=4, value=msg_type or '')
            ws.cell(row=row_num, column=5, value=channel or '')
            ws.cell(row=row_num, column=6, value=subject or '')
            ws.cell(row=row_num, column=7, value=score_effect or 0)
        
        # ضبط عرض الأعمدة
        column_widths = [6, 25, 12, 12, 12, 40, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # تجميد الصف الأول
        ws.freeze_panes = 'A2'
        
        # حفظ الملف
        wb.save(file_path)
        return True
        
    except Exception as e:
        print(f"خطأ في تصدير الرسائل Excel: {e}")
        return False


def export_requests_to_csv(file_path: str) -> bool:
    """
    تصدير الطلبات إلى ملف CSV
    
    Args:
        file_path: مسار ملف CSV
    
    Returns:
        True في حالة النجاح، False في حالة الفشل
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        query = """
            SELECT 
                r.id, c.company_name, r.client_email, r.request_type,
                r.status, r.reply_status, r.created_at
            FROM requests r
            LEFT JOIN clients c ON r.client_id = c.id
            ORDER BY r.id DESC
        """
        cur.execute(query)
        requests = cur.fetchall()
        conn.close()
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # رأس الجدول
            writer.writerow([
                'ID', 'Client', 'Email', 'Request Type', 'Status', 'Reply Status', 'Created At'
            ])
            
            # البيانات
            for req in requests:
                req_id, company, email, req_type, status, reply_status, created_at = req
                
                writer.writerow([
                    req_id, company or '', email or '', req_type or '',
                    status or '', reply_status or '', created_at or ''
                ])
        
        return True
        
    except Exception as e:
        print(f"خطأ في تصدير الطلبات CSV: {e}")
        return False


def export_requests_to_excel(file_path: str) -> bool:
    """
    تصدير الطلبات إلى ملف Excel
    
    Args:
        file_path: مسار ملف Excel (.xlsx)
    
    Returns:
        True في حالة النجاح، False في حالة الفشل
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install it using: pip install openpyxl")
    
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        query = """
            SELECT 
                r.id, c.company_name, r.client_email, r.request_type,
                r.status, r.reply_status, r.created_at
            FROM requests r
            LEFT JOIN clients c ON r.client_id = c.id
            ORDER BY r.id DESC
        """
        cur.execute(query)
        requests = cur.fetchall()
        conn.close()
        
        # إنشاء workbook جديد
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requests"
        
        # تنسيق الرأس
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # رأس الجدول
        headers = ['ID', 'Client', 'Email', 'Request Type', 'Status', 'Reply Status', 'Created At']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # البيانات
        for row_num, req in enumerate(requests, 2):
            req_id, company, email, req_type, status, reply_status, created_at = req
            
            ws.cell(row=row_num, column=1, value=req_id)
            ws.cell(row=row_num, column=2, value=company or '')
            ws.cell(row=row_num, column=3, value=email or '')
            ws.cell(row=row_num, column=4, value=req_type or '')
            ws.cell(row=row_num, column=5, value=status or '')
            ws.cell(row=row_num, column=6, value=reply_status or '')
            ws.cell(row=row_num, column=7, value=created_at or '')
        
        # ضبط عرض الأعمدة
        column_widths = [6, 25, 30, 20, 12, 15, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # تجميد الصف الأول
        ws.freeze_panes = 'A2'
        
        # حفظ الملف
        wb.save(file_path)
        return True
        
    except Exception as e:
        print(f"خطأ في تصدير الطلبات Excel: {e}")
        return False


def export_full_report_to_excel(file_path: str) -> bool:
    """
    تصدير تقرير شامل يحتوي على جميع البيانات في ملف Excel واحد
    
    Args:
        file_path: مسار ملف Excel (.xlsx)
    
    Returns:
        True في حالة النجاح، False في حالة الفشل
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install it using: pip install openpyxl")
    
    try:
        wb = openpyxl.Workbook()
        
        # حذف الورقة الافتراضية
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 1. ورقة العملاء
        clients = get_all_clients()
        ws_clients = wb.create_sheet("Clients")
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        headers = [
            'ID', 'Company Name', 'Country', 'Contact Person',
            'Email', 'Phone', 'Website', 'Date Added',
            'Status', 'Score', 'Classification', 'Focus'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_clients.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_num, client in enumerate(clients, 2):
            (
                client_id, company, country, contact, email,
                phone, website, date_added, status, score,
                classification, is_focus
            ) = client
            
            ws_clients.cell(row=row_num, column=1, value=client_id)
            ws_clients.cell(row=row_num, column=2, value=company or '')
            ws_clients.cell(row=row_num, column=3, value=country or '')
            ws_clients.cell(row=row_num, column=4, value=contact or '')
            ws_clients.cell(row=row_num, column=5, value=email or '')
            ws_clients.cell(row=row_num, column=6, value=phone or '')
            ws_clients.cell(row=row_num, column=7, value=website or '')
            ws_clients.cell(row=row_num, column=8, value=date_added or '')
            ws_clients.cell(row=row_num, column=9, value=status or '')
            ws_clients.cell(row=row_num, column=10, value=score or 0)
            ws_clients.cell(row=row_num, column=11, value=classification or '')
            ws_clients.cell(row=row_num, column=12, value='Yes' if is_focus else 'No')
        
        column_widths = [6, 25, 15, 20, 30, 15, 25, 12, 18, 8, 20, 8]
        for col_num, width in enumerate(column_widths, 1):
            ws_clients.column_dimensions[get_column_letter(col_num)].width = width
        ws_clients.freeze_panes = 'A2'
        
        # 2. ورقة الرسائل
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT 
                m.id, c.company_name, m.message_date, m.message_type,
                m.channel, m.client_response, m.score_effect
            FROM messages m
            JOIN clients c ON m.client_id = c.id
            ORDER BY m.message_date DESC
        """)
        messages = cur.fetchall()
        
        ws_messages = wb.create_sheet("Messages")
        
        header_fill2 = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        headers2 = ['ID', 'Client', 'Date', 'Type', 'Channel', 'Subject', 'Score Effect']
        
        for col_num, header in enumerate(headers2, 1):
            cell = ws_messages.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill2
            cell.font = header_font
        
        for row_num, msg in enumerate(messages, 2):
            msg_id, company, msg_date, msg_type, channel, subject, score_effect = msg
            ws_messages.cell(row=row_num, column=1, value=msg_id)
            ws_messages.cell(row=row_num, column=2, value=company or '')
            ws_messages.cell(row=row_num, column=3, value=msg_date or '')
            ws_messages.cell(row=row_num, column=4, value=msg_type or '')
            ws_messages.cell(row=row_num, column=5, value=channel or '')
            ws_messages.cell(row=row_num, column=6, value=subject or '')
            ws_messages.cell(row=row_num, column=7, value=score_effect or 0)
        
        column_widths2 = [6, 25, 12, 12, 12, 40, 12]
        for col_num, width in enumerate(column_widths2, 1):
            ws_messages.column_dimensions[get_column_letter(col_num)].width = width
        ws_messages.freeze_panes = 'A2'
        
        # 3. ورقة الطلبات
        cur.execute("""
            SELECT 
                r.id, c.company_name, r.client_email, r.request_type,
                r.status, r.reply_status, r.created_at
            FROM requests r
            LEFT JOIN clients c ON r.client_id = c.id
            ORDER BY r.id DESC
        """)
        requests = cur.fetchall()
        conn.close()
        
        ws_requests = wb.create_sheet("Requests")
        
        header_fill3 = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        headers3 = ['ID', 'Client', 'Email', 'Request Type', 'Status', 'Reply Status', 'Created At']
        
        for col_num, header in enumerate(headers3, 1):
            cell = ws_requests.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill3
            cell.font = header_font
        
        for row_num, req in enumerate(requests, 2):
            req_id, company, email, req_type, status, reply_status, created_at = req
            ws_requests.cell(row=row_num, column=1, value=req_id)
            ws_requests.cell(row=row_num, column=2, value=company or '')
            ws_requests.cell(row=row_num, column=3, value=email or '')
            ws_requests.cell(row=row_num, column=4, value=req_type or '')
            ws_requests.cell(row=row_num, column=5, value=status or '')
            ws_requests.cell(row=row_num, column=6, value=reply_status or '')
            ws_requests.cell(row=row_num, column=7, value=created_at or '')
        
        column_widths3 = [6, 25, 30, 20, 12, 15, 18]
        for col_num, width in enumerate(column_widths3, 1):
            ws_requests.column_dimensions[get_column_letter(col_num)].width = width
        ws_requests.freeze_panes = 'A2'
        
        # 4. ورقة الإحصائيات
        ws_stats = wb.create_sheet("Statistics")
        
        stats_headers = ['Metric', 'Value']
        for col_num, header in enumerate(stats_headers, 1):
            cell = ws_stats.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # حساب الإحصائيات
        total_clients = len(clients)
        serious_count = sum(1 for c in clients if c[10] and '🔥' in str(c[10]))
        potential_count = sum(1 for c in clients if c[10] and '👍' in str(c[10]))
        focus_count = sum(1 for c in clients if c[11] == 1)
        total_messages = len(messages)
        total_requests = len(requests)
        
        stats_data = [
            ('Total Clients', total_clients),
            ('Serious Buyers', serious_count),
            ('Potential Clients', potential_count),
            ('Focus Clients', focus_count),
            ('Total Messages', total_messages),
            ('Total Requests', total_requests),
            ('Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        for row_num, (metric, value) in enumerate(stats_data, 2):
            ws_stats.cell(row=row_num, column=1, value=metric)
            ws_stats.cell(row=row_num, column=2, value=value)
        
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 20
        
        # حفظ الملف
        wb.save(file_path)
        return True
        
    except Exception as e:
        print(f"خطأ في تصدير التقرير الشامل: {e}")
        return False
