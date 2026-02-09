"""
نظام استيراد البيانات من CSV و Excel
Data Import System from CSV and Excel
"""
import csv
import os
from typing import List, Dict, Optional, Tuple
from datetime import datetime

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    openpyxl = None

from .db import (
    get_connection, add_client, add_message,
    find_client_by_email, get_client_by_id
)
from .sales import add_sale_deal, init_sales_table
from .validation import validate_email, validate_phone, validate_url


def import_clients_from_csv(file_path: str) -> Dict[str, int]:
    """
    استيراد العملاء من ملف CSV
    
    Args:
        file_path: مسار ملف CSV
    
    Returns:
        Dict containing: {'success': count, 'failed': count, 'skipped': count, 'errors': list}
    """
    results = {
        'success': 0,
        'failed': 0,
        'skipped': 0,
        'errors': []
    }
    
    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            # محاولة اكتشاف الفواصل تلقائياً
            sample = f.read(1024)
            f.seek(0)
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter
            
            reader = csv.DictReader(f, delimiter=delimiter)
            
            # قراءة الصف الأول كرؤوس
            headers = reader.fieldnames
            if not headers:
                results['errors'].append("الملف فارغ أو لا يحتوي على رؤوس")
                return results
            
            # تحويل الرؤوس إلى lowercase للتوافق
            headers_lower = [h.lower().strip() if h else '' for h in headers]
            
            for row_num, row in enumerate(reader, start=2):  # start=2 لأن الصف الأول هو الرؤوس
                try:
                    # إنشاء قاموس للبيانات
                    client_data = {}
                    
                    # البحث عن الأعمدة المطلوبة
                    for i, header in enumerate(headers):
                        header_lower = headers_lower[i]
                        value = row.get(header, '').strip() if header else ''
                        
                        # تعيين القيم حسب اسم العمود
                        if 'company' in header_lower or 'اسم' in header_lower or 'name' in header_lower:
                            client_data['company_name'] = value
                        elif 'email' in header_lower or 'بريد' in header_lower:
                            client_data['email'] = value
                        elif 'phone' in header_lower or 'هاتف' in header_lower or 'tel' in header_lower:
                            client_data['phone'] = value
                        elif 'country' in header_lower or 'بلد' in header_lower:
                            client_data['country'] = value
                        elif 'website' in header_lower or 'موقع' in header_lower or 'url' in header_lower:
                            client_data['website'] = value
                        elif 'classification' in header_lower or 'تصنيف' in header_lower:
                            client_data['classification'] = value
                        elif 'score' in header_lower or 'نقاط' in header_lower:
                            try:
                                client_data['score'] = float(value) if value else 0
                            except:
                                client_data['score'] = 0
                    
                    # التحقق من البيانات الأساسية
                    if not client_data.get('company_name'):
                        results['skipped'] += 1
                        results['errors'].append(f"الصف {row_num}: اسم الشركة مطلوب")
                        continue
                    
                    # التحقق من البريد الإلكتروني
                    email = client_data.get('email', '')
                    if email:
                        if not validate_email(email):
                            results['skipped'] += 1
                            results['errors'].append(f"الصف {row_num}: بريد إلكتروني غير صحيح: {email}")
                            continue
                        
                        # التحقق من وجود العميل
                        existing = find_client_by_email(email)
                        if existing:
                            results['skipped'] += 1
                            results['errors'].append(f"الصف {row_num}: العميل موجود بالفعل: {email}")
                            continue
                    
                    # التحقق من الهاتف
                    phone = client_data.get('phone', '')
                    if phone and not validate_phone(phone):
                        # تحذير فقط، لا نوقف الاستيراد
                        pass
                    
                    # التحقق من الموقع
                    website = client_data.get('website', '')
                    if website and not validate_url(website):
                        # تحذير فقط
                        pass
                    
                    # إضافة العميل
                    client_id = add_client(client_data)
                    if client_id:
                        results['success'] += 1
                    else:
                        results['failed'] += 1
                        results['errors'].append(f"الصف {row_num}: فشل إضافة العميل")
                
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"الصف {row_num}: {str(e)}")
    
    except Exception as e:
        results['errors'].append(f"خطأ في قراءة الملف: {str(e)}")
    
    return results


def import_clients_from_excel(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, int]:
    """
    استيراد العملاء من ملف Excel
    
    Args:
        file_path: مسار ملف Excel
        sheet_name: اسم الورقة (إذا كان None، يتم استخدام الورقة الأولى)
    
    Returns:
        Dict containing: {'success': count, 'failed': count, 'skipped': count, 'errors': list}
    """
    if not EXCEL_AVAILABLE:
        return {
            'success': 0,
            'failed': 0,
            'skipped': 0,
            'errors': ['openpyxl غير مثبت. قم بتثبيته باستخدام: pip install openpyxl']
        }
    
    results = {
        'success': 0,
        'failed': 0,
        'skipped': 0,
        'errors': []
    }
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # اختيار الورقة
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                results['errors'].append(f"الورقة '{sheet_name}' غير موجودة")
                return results
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        # قراءة الرؤوس من الصف الأول
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value if cell.value else '')
        
        if not headers or all(not h for h in headers):
            results['errors'].append("الملف فارغ أو لا يحتوي على رؤوس")
            return results
        
        # تحويل الرؤوس إلى lowercase
        headers_lower = [str(h).lower().strip() if h else '' for h in headers]
        
        # قراءة البيانات
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            try:
                # إنشاء قاموس للبيانات
                client_data = {}
                
                # استخراج القيم
                for i, cell in enumerate(row):
                    if i >= len(headers):
                        break
                    
                    header_lower = headers_lower[i]
                    value = str(cell.value).strip() if cell.value else ''
                    
                    # تعيين القيم حسب اسم العمود
                    if 'company' in header_lower or 'اسم' in header_lower or 'name' in header_lower:
                        client_data['company_name'] = value
                    elif 'email' in header_lower or 'بريد' in header_lower:
                        client_data['email'] = value
                    elif 'phone' in header_lower or 'هاتف' in header_lower or 'tel' in header_lower:
                        client_data['phone'] = value
                    elif 'country' in header_lower or 'بلد' in header_lower:
                        client_data['country'] = value
                    elif 'website' in header_lower or 'موقع' in header_lower or 'url' in header_lower:
                        client_data['website'] = value
                    elif 'classification' in header_lower or 'تصنيف' in header_lower:
                        client_data['classification'] = value
                    elif 'score' in header_lower or 'نقاط' in header_lower:
                        try:
                            client_data['score'] = float(value) if value else 0
                        except:
                            client_data['score'] = 0
                
                # التحقق من البيانات الأساسية
                if not client_data.get('company_name'):
                    results['skipped'] += 1
                    results['errors'].append(f"الصف {row_num}: اسم الشركة مطلوب")
                    continue
                
                # التحقق من البريد الإلكتروني
                email = client_data.get('email', '')
                if email:
                    if not validate_email(email):
                        results['skipped'] += 1
                        results['errors'].append(f"الصف {row_num}: بريد إلكتروني غير صحيح: {email}")
                        continue
                    
                    # التحقق من وجود العميل
                    existing = find_client_by_email(email)
                    if existing:
                        results['skipped'] += 1
                        results['errors'].append(f"الصف {row_num}: العميل موجود بالفعل: {email}")
                        continue
                
                # إضافة العميل
                client_id = add_client(client_data)
                if client_id:
                    results['success'] += 1
                else:
                    results['failed'] += 1
                    results['errors'].append(f"الصف {row_num}: فشل إضافة العميل")
            
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"الصف {row_num}: {str(e)}")
        
        workbook.close()
    
    except Exception as e:
        results['errors'].append(f"خطأ في قراءة الملف: {str(e)}")
    
    return results


def import_messages_from_csv(file_path: str) -> Dict[str, int]:
    """
    استيراد الرسائل من ملف CSV
    
    Args:
        file_path: مسار ملف CSV
    
    Returns:
        Dict containing: {'success': count, 'failed': count, 'skipped': count, 'errors': list}
    """
    results = {
        'success': 0,
        'failed': 0,
        'skipped': 0,
        'errors': []
    }
    
    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            sample = f.read(1024)
            f.seek(0)
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter
            
            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames
            
            if not headers:
                results['errors'].append("الملف فارغ أو لا يحتوي على رؤوس")
                return results
            
            headers_lower = [h.lower().strip() if h else '' for h in headers]
            
            for row_num, row in enumerate(reader, start=2):
                try:
                    message_data = {}
                    client_id = None
                    
                    # استخراج البيانات
                    for i, header in enumerate(headers):
                        header_lower = headers_lower[i]
                        value = row.get(header, '').strip() if header else ''
                        
                        if 'client' in header_lower or 'عميل' in header_lower:
                            # يمكن أن يكون ID أو email أو company name
                            if value.isdigit():
                                client_id = int(value)
                            else:
                                # البحث بالبريد أو الاسم
                                if '@' in value:
                                    client = find_client_by_email(value)
                                    if client:
                                        client_id = client[0]
                                else:
                                    # البحث بالاسم
                                    from .db import get_all_clients
                                    clients = get_all_clients()
                                    for c in clients:
                                        if c.get('company_name', '').lower() == value.lower():
                                            client_id = c.get('id')
                                            break
                        
                        elif 'email' in header_lower or 'بريد' in header_lower:
                            message_data['sender_email'] = value
                        elif 'subject' in header_lower or 'موضوع' in header_lower:
                            message_data['subject'] = value
                        elif 'body' in header_lower or 'content' in header_lower or 'محتوى' in header_lower:
                            message_data['body'] = value
                        elif 'date' in header_lower or 'تاريخ' in header_lower:
                            message_data['date'] = value
                        elif 'channel' in header_lower or 'قناة' in header_lower:
                            message_data['channel'] = value
                    
                    # التحقق من العميل
                    if not client_id:
                        results['skipped'] += 1
                        results['errors'].append(f"الصف {row_num}: لم يتم العثور على العميل")
                        continue
                    
                    # التحقق من البيانات الأساسية
                    body = message_data.get('body', '')
                    subject = message_data.get('subject', '')
                    if not subject and not body:
                        results['skipped'] += 1
                        results['errors'].append(f"الصف {row_num}: الموضوع أو المحتوى مطلوب")
                        continue
                    
                    # إضافة الرسالة
                    # تحضير البيانات للرسالة
                    message_date = message_data.get('date')
                    if not message_date:
                        message_date = datetime.now().strftime("%d/%m/%Y")
                    
                    # تحديد نوع الرسالة
                    message_type = "inquiry"
                    body_lower = body.lower() if body else ''
                    subject_lower = subject.lower() if subject else ''
                    content_lower = body_lower + ' ' + subject_lower
                    
                    if any(word in content_lower for word in ['price', 'cost', 'سعر', 'تكلفة']):
                        message_type = "price_request"
                    elif any(word in content_lower for word in ['sample', 'عينة']):
                        message_type = "sample_request"
                    elif any(word in content_lower for word in ['spec', 'specification', 'مواصفات']):
                        message_type = "spec_request"
                    elif any(word in content_lower for word in ['moq', 'minimum', 'أقل', 'حد أدنى']):
                        message_type = "moq_request"
                    
                    # حساب تأثير النقاط
                    score_effect = 0
                    if message_type in ["price_request", "sample_request", "spec_request", "moq_request"]:
                        score_effect = 5  # طلب = +5 نقاط
                    
                    message_id = add_message({
                        'client_id': client_id,
                        'message_date': message_date,
                        'message_type': message_type,
                        'channel': message_data.get('channel', 'email'),
                        'client_response': body,
                        'notes': subject,
                        'score_effect': score_effect
                    })
                    
                    if message_id:
                        results['success'] += 1
                    else:
                        results['failed'] += 1
                        results['errors'].append(f"الصف {row_num}: فشل إضافة الرسالة")
                
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"الصف {row_num}: {str(e)}")
    
    except Exception as e:
        results['errors'].append(f"خطأ في قراءة الملف: {str(e)}")
    
    return results


def import_deals_from_csv(file_path: str) -> Dict[str, int]:
    """
    استيراد الصفقات من ملف CSV
    
    Args:
        file_path: مسار ملف CSV
    
    Returns:
        Dict containing: {'success': count, 'failed': count, 'skipped': count, 'errors': list}
    """
    results = {
        'success': 0,
        'failed': 0,
        'skipped': 0,
        'errors': []
    }
    
    try:
        init_sales_table()
        
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            sample = f.read(1024)
            f.seek(0)
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter
            
            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames
            
            if not headers:
                results['errors'].append("الملف فارغ أو لا يحتوي على رؤوس")
                return results
            
            headers_lower = [h.lower().strip() if h else '' for h in headers]
            
            for row_num, row in enumerate(reader, start=2):
                try:
                    deal_data = {}
                    client_id = None
                    
                    # استخراج البيانات
                    for i, header in enumerate(headers):
                        header_lower = headers_lower[i]
                        value = row.get(header, '').strip() if header else ''
                        
                        if 'client' in header_lower or 'عميل' in header_lower:
                            if value.isdigit():
                                client_id = int(value)
                            else:
                                if '@' in value:
                                    client = find_client_by_email(value)
                                    if client:
                                        client_id = client[0]
                        
                        elif 'deal' in header_lower or 'صفقة' in header_lower or 'name' in header_lower:
                            deal_data['deal_name'] = value
                        elif 'product' in header_lower or 'منتج' in header_lower:
                            deal_data['product_name'] = value
                        elif 'stage' in header_lower or 'مرحلة' in header_lower:
                            deal_data['stage'] = value
                        elif 'value' in header_lower or 'قيمة' in header_lower:
                            try:
                                deal_data['value'] = float(value) if value else 0
                            except:
                                deal_data['value'] = 0
                        elif 'currency' in header_lower or 'عملة' in header_lower:
                            deal_data['currency'] = value or 'USD'
                        elif 'probability' in header_lower or 'احتمال' in header_lower:
                            try:
                                deal_data['probability'] = float(value) if value else 0.1
                            except:
                                deal_data['probability'] = 0.1
                        elif 'expected' in header_lower or 'متوقع' in header_lower or 'close' in header_lower:
                            deal_data['expected_close_date'] = value
                    
                    # التحقق من العميل
                    if not client_id:
                        results['skipped'] += 1
                        results['errors'].append(f"الصف {row_num}: لم يتم العثور على العميل")
                        continue
                    
                    # التحقق من البيانات الأساسية
                    if not deal_data.get('deal_name'):
                        results['skipped'] += 1
                        results['errors'].append(f"الصف {row_num}: اسم الصفقة مطلوب")
                        continue
                    
                    # إضافة الصفقة
                    deal_id = add_sale_deal({
                        'client_id': client_id,
                        **deal_data
                    })
                    
                    if deal_id:
                        results['success'] += 1
                    else:
                        results['failed'] += 1
                        results['errors'].append(f"الصف {row_num}: فشل إضافة الصفقة")
                
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"الصف {row_num}: {str(e)}")
    
    except Exception as e:
        results['errors'].append(f"خطأ في قراءة الملف: {str(e)}")
    
    return results
