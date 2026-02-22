"""
نافذة البحث المتخصصة عن مستوردي البصل والكراث المجفف
Specialized Search Window for Dried Onion and Leek Importers
"""
from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                             QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
                             QMessageBox, QProgressBar, QTextEdit, QHeaderView, QComboBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont
from typing import List, Dict, Optional
from core.specialized_importer_search import search_dried_onion_leek_importers


class SpecializedSearchThread(QThread):
    """Thread للبحث في الخلفية"""
    finished = pyqtSignal(list)
    error = pyqtSignal(str)
    progress = pyqtSignal(str)
    
    def __init__(self, api_key: str, max_results: int = 50, api_provider: str = "serper", 
                 product_name: str = "", country: str = "USA"):
        super().__init__()
        self.api_key = api_key
        self.max_results = max_results
        self.api_provider = api_provider
        self.product_name = product_name
        self.country = country
    
    def run(self):
        try:
            provider_name = "Serper.dev" if self.api_provider == "serper" else "SerpAPI"
            self.progress.emit(f"بدء البحث المتخصص باستخدام {provider_name}...")
            print(f"DEBUG: بدء البحث - Provider: {provider_name}, Product: {self.product_name}, Country: {self.country}, Max Results: {self.max_results}")
            results = search_dried_onion_leek_importers(
                self.api_key, 
                self.max_results, 
                self.api_provider,
                self.product_name,
                self.country
            )
            print(f"DEBUG: انتهى البحث - عدد النتائج: {len(results)}")
            self.finished.emit(results)
        except Exception as e:
            import traceback
            error_msg = f"{str(e)}\n\n{traceback.format_exc()}"
            print(f"DEBUG: خطأ في Thread: {error_msg}")
            self.error.emit(str(e))


class SpecializedSearchWindow(QDialog):
    """نافذة البحث المتخصصة"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🔍 البحث المتخصص - مستوردي البصل والكراث والسبانخ المجفف من مصر")
        self.setMinimumSize(1000, 700)
        self.results = []
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # العنوان
        title = QLabel("🔍 البحث المتخصص عن مستوردي البصل والكراث والسبانخ المجفف من مصر")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # وصف
        description = QLabel(
            "هذا البرنامج متخصص في البحث عن شركات حقيقية تستورد:\n"
            "• بصل مجفف (Dehydrated/Dried Onion)\n"
            "• كراث مجفف (Dehydrated/Dried Leek)\n"
            "• سبانخ مجفف (Dehydrated/Dried Spinach)\n"
            "من مصر فقط - بدون شركات مصرية أو صينية أو هندية"
        )
        description.setAlignment(Qt.AlignCenter)
        description.setStyleSheet("color: #666; padding: 10px;")
        layout.addWidget(description)
        
        # اختيار نوع API
        api_type_layout = QHBoxLayout()
        api_type_layout.addWidget(QLabel("نوع API:"))
        self.api_type_combo = QComboBox()
        self.api_type_combo.addItems(["Serper.dev", "SerpAPI"])
        self.api_type_combo.setCurrentIndex(0)  # Serper.dev كافتراضي
        api_type_layout.addWidget(self.api_type_combo)
        api_type_layout.addStretch()
        layout.addLayout(api_type_layout)
        
        # حقل مفتاح API
        api_layout = QHBoxLayout()
        api_layout.addWidget(QLabel("مفتاح API:"))
        self.api_key_input = QLineEdit()
        self.api_key_input.setPlaceholderText("أدخل مفتاح API الخاص بك...")
        self.api_key_input.setEchoMode(QLineEdit.Password)
        api_layout.addWidget(self.api_key_input)
        
        # زر إظهار/إخفاء المفتاح
        self.show_key_btn = QPushButton("👁️")
        self.show_key_btn.setMaximumWidth(40)
        self.show_key_btn.clicked.connect(self.toggle_key_visibility)
        api_layout.addWidget(self.show_key_btn)
        
        layout.addLayout(api_layout)
        
        # حقل اسم المنتج
        product_layout = QHBoxLayout()
        product_layout.addWidget(QLabel("اسم المنتج (اختياري):"))
        self.product_input = QLineEdit()
        self.product_input.setPlaceholderText("مثال: dehydrated onion, dried leek, spinach...")
        product_layout.addWidget(self.product_input)
        layout.addLayout(product_layout)
        
        # اختيار البلد
        country_layout = QHBoxLayout()
        country_layout.addWidget(QLabel("البلد:"))
        self.country_combo = QComboBox()
        self.country_combo.addItems([
            "USA",
            "United Kingdom",
            "Germany",
            "France",
            "Italy",
            "Spain",
            "Netherlands",
            "Belgium",
            "Switzerland",
            "Canada",
            "Australia",
            "Japan",
            "South Korea",
            "All Countries"  # خيار للبحث في جميع البلدان
        ])
        self.country_combo.setCurrentIndex(0)  # USA كافتراضي
        country_layout.addWidget(self.country_combo)
        country_layout.addStretch()
        layout.addLayout(country_layout)
        
        # عدد النتائج
        results_layout = QHBoxLayout()
        results_layout.addWidget(QLabel("الحد الأقصى للنتائج:"))
        self.max_results_input = QLineEdit("50")
        self.max_results_input.setMaximumWidth(100)
        results_layout.addWidget(self.max_results_input)
        results_layout.addStretch()
        layout.addLayout(results_layout)
        
        # أزرار
        buttons_layout = QHBoxLayout()
        self.search_btn = QPushButton("🔍 بدء البحث المتخصص")
        self.search_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px; font-size: 12px;")
        self.search_btn.clicked.connect(self.start_search)
        buttons_layout.addWidget(self.search_btn)
        
        self.export_btn = QPushButton("📥 تصدير النتائج")
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_results)
        buttons_layout.addWidget(self.export_btn)
        
        layout.addLayout(buttons_layout)
        
        # شريط التقدم
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # رسائل التقدم
        self.progress_text = QTextEdit()
        self.progress_text.setMaximumHeight(100)
        self.progress_text.setReadOnly(True)
        self.progress_text.setVisible(False)
        layout.addWidget(self.progress_text)
        
        # جدول النتائج
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(6)
        self.results_table.setHorizontalHeaderLabels([
            "اسم الشركة", "الموقع", "البريد الإلكتروني", 
            "الهاتف", "الدولة", "المقتطف"
        ])
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.results_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.results_table)
        
        # إحصائيات
        self.stats_label = QLabel("لم يتم البحث بعد")
        self.stats_label.setAlignment(Qt.AlignCenter)
        self.stats_label.setStyleSheet("font-weight: bold; padding: 5px;")
        layout.addWidget(self.stats_label)
        
        self.setLayout(layout)
    
    def toggle_key_visibility(self):
        """إظهار/إخفاء مفتاح API"""
        if self.api_key_input.echoMode() == QLineEdit.Password:
            self.api_key_input.setEchoMode(QLineEdit.Normal)
            self.show_key_btn.setText("🙈")
        else:
            self.api_key_input.setEchoMode(QLineEdit.Password)
            self.show_key_btn.setText("👁️")
    
    def start_search(self):
        """بدء البحث"""
        try:
            api_key = self.api_key_input.text().strip()
            if not api_key:
                QMessageBox.warning(self, "تحذير", "يرجى إدخال مفتاح API")
                return
            
            try:
                max_results = int(self.max_results_input.text())
            except:
                max_results = 50
            
            # إعداد الواجهة
            self.search_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_text.setVisible(True)
            self.progress_bar.setRange(0, 0)  # indeterminate
            self.progress_text.clear()
            self.results_table.setRowCount(0)
            self.results = []
            
            # تحديد نوع API
            api_provider = "serper" if self.api_type_combo.currentText() == "Serper.dev" else "serpapi"
            
            # الحصول على اسم المنتج والبلد
            product_name = self.product_input.text().strip()
            country = self.country_combo.currentText()
            
            # بدء البحث في thread منفصل
            self.search_thread = SpecializedSearchThread(api_key, max_results, api_provider, product_name, country)
            self.search_thread.progress.connect(self.update_progress)
            self.search_thread.finished.connect(self.on_search_finished)
            self.search_thread.error.connect(self.on_search_error)
            self.search_thread.start()
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ عند بدء البحث:\n{str(e)}")
            self.search_btn.setEnabled(True)
            self.progress_bar.setVisible(False)
            self.progress_text.setVisible(False)
    
    def update_progress(self, message: str):
        """تحديث رسالة التقدم"""
        self.progress_text.append(message)
    
    def on_search_finished(self, results: List[dict]):
        """عند انتهاء البحث"""
        self.search_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.progress_text.setVisible(False)
        
        self.results = results
        self.display_results(results)
        
        if results:
            self.export_btn.setEnabled(True)
            QMessageBox.information(
                self, "نجح البحث", 
                f"تم العثور على {len(results)} شركة حقيقية"
            )
        else:
            QMessageBox.warning(
                self, "لا توجد نتائج", 
                "لم يتم العثور على شركات حقيقية.\n"
                "يرجى التحقق من مفتاح SerpAPI أو المحاولة مرة أخرى."
            )
    
    def on_search_error(self, error: str):
        """عند حدوث خطأ"""
        self.search_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.progress_text.setVisible(False)
        QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء البحث:\n{error}")
    
    def display_results(self, results: List[dict]):
        """عرض النتائج في الجدول"""
        self.results_table.setRowCount(len(results))
        
        for row, result in enumerate(results):
            self.results_table.setItem(row, 0, QTableWidgetItem(result.get("company_name", "")))
            self.results_table.setItem(row, 1, QTableWidgetItem(result.get("website", "")))
            self.results_table.setItem(row, 2, QTableWidgetItem(result.get("email", "")))
            self.results_table.setItem(row, 3, QTableWidgetItem(result.get("phone", "")))
            self.results_table.setItem(row, 4, QTableWidgetItem(result.get("country", "")))
            self.results_table.setItem(row, 5, QTableWidgetItem(result.get("snippet", "")))
        
        # ضبط عرض الأعمدة
        self.results_table.resizeColumnsToContents()
        
        # تحديث الإحصائيات
        self.stats_label.setText(
            f"تم العثور على {len(results)} شركة حقيقية | "
            f"مع بريد إلكتروني: {sum(1 for r in results if r.get('email'))} | "
            f"مع هاتف: {sum(1 for r in results if r.get('phone'))}"
        )
    
    def export_results(self):
        """تصدير النتائج إلى Excel"""
        if not self.results:
            QMessageBox.warning(self, "تحذير", "لا توجد نتائج للتصدير")
            return
        
        from PyQt5.QtWidgets import QFileDialog
        import os
        from datetime import datetime
        
        default_filename = f"dried_onion_leek_importers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "حفظ النتائج", default_filename,
            "Excel Files (*.xlsx);;CSV Files (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            if file_path.endswith('.xlsx'):
                self.export_to_excel(file_path)
            else:
                self.export_to_csv(file_path)
            
            QMessageBox.information(self, "نجح التصدير", f"تم تصدير {len(self.results)} شركة إلى:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل التصدير:\n{str(e)}")
    
    def export_to_excel(self, file_path: str):
        """تصدير إلى Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "مستوردي البصل والكراث المجفف"
        
        # رؤوس الأعمدة
        headers = ["اسم الشركة", "الموقع", "البريد الإلكتروني", "الهاتف", "الدولة", "المقتطف"]
        ws.append(headers)
        
        # تنسيق الرؤوس
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # البيانات
        for result in self.results:
            ws.append([
                result.get("company_name", ""),
                result.get("website", ""),
                result.get("email", ""),
                result.get("phone", ""),
                result.get("country", ""),
                result.get("snippet", "")
            ])
        
        # ضبط عرض الأعمدة
        column_widths = [30, 40, 25, 20, 15, 50]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
        
        wb.save(file_path)
    
    def export_to_csv(self, file_path: str):
        """تصدير إلى CSV"""
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(["اسم الشركة", "الموقع", "البريد الإلكتروني", "الهاتف", "الدولة", "المقتطف"])
            
            for result in self.results:
                writer.writerow([
                    result.get("company_name", ""),
                    result.get("website", ""),
                    result.get("email", ""),
                    result.get("phone", ""),
                    result.get("country", ""),
                    result.get("snippet", "")
                ])
